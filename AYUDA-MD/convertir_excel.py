import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ── Rutas ──────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
IN_PATH    = os.path.join(SCRIPT_DIR, "001.xlsx")
OUT_PATH   = os.path.join(SCRIPT_DIR, "catalogo_importacion.xlsx")

if not os.path.exists(IN_PATH):
    print(f"Error: No se encontró el archivo de origen: {IN_PATH}")
    sys.exit(1)

# ── Helpers ────────────────────────────────────────────────────────────────────
def guess_unit(name):
    name_upper = name.upper()
    if ' LT' in name_upper or 'LITRO' in name_upper: return 'Litro'
    if ' ML' in name_upper or 'MILILITRO' in name_upper: return 'Mililitro'
    if ' KG' in name_upper or 'KILO' in name_upper: return 'Kilogramo'
    if ' GR' in name_upper or 'GRAMO' in name_upper: return 'Gramo'
    if ' CAJA' in name_upper or ' CJ' in name_upper: return 'Caja'
    if ' DOCENA' in name_upper or ' DOC ' in name_upper: return 'Docena'
    if ' LATA' in name_upper: return 'Lata'
    if ' PAQUETE' in name_upper or ' PQT' in name_upper: return 'Paquete'
    if ' BOLSA' in name_upper: return 'Bolsa'
    return 'Unidad'

# ── 1. Leer archivo original (001.xlsx) ────────────────────────────────────────
print(f"Leyendo archivo origen: {IN_PATH} ...")
try:
    xls = pd.ExcelFile(IN_PATH)
except Exception as e:
    print(f"Error al leer el Excel: {e}")
    sys.exit(1)

products = []

for sheet in xls.sheet_names:
    if sheet.lower() in ["hoja1", "hoja2", "hoja3"]:
        continue
    
    try:
        df = pd.read_excel(xls, sheet_name=sheet, header=None)
    except Exception:
        continue
    
    if len(df) <= 2:
        continue
    
    category_name = sheet.strip().upper()
    
    for i in range(2, len(df)):
        row = df.iloc[i]
        
        prod_name = str(row[0]).strip()
        if pd.isna(row[0]) or not prod_name or prod_name.lower() == "nan":
            continue
            
        try:
            price = float(row[1]) if not pd.isna(row[1]) else 0.0
        except ValueError:
            price = 0.0
                
        try:
            stock = float(row[2]) if not pd.isna(row[2]) else 0.0
        except ValueError:
            stock = 0.0

        products.append({
            "nombre":        prod_name,
            "categoria":     category_name,
            "unidad":        guess_unit(prod_name),
            "stock":         round(stock, 2),
            "precio_compra": round(price * 0.7, 2),  # Margen estimado 30%
            "precio_venta":  round(price, 2),
        })

print(f"✓ Se extrajeron {len(products)} productos en total.")

# Agrupar por categoría
by_cat = defaultdict(list)
for p in products:
    by_cat[p["categoria"]].append(p)

# ── 2. Generar el nuevo Excel ──────────────────────────────────────────────────
print("Generando nuevo formato con estilos...")

HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT    = Font(name="Calibri", size=10)
DATA_ALIGN_L = Alignment(horizontal="left",   vertical="center")
DATA_ALIGN_C = Alignment(horizontal="center", vertical="center")
DATA_ALIGN_R = Alignment(horizontal="right",  vertical="center")

ALT_FILL     = PatternFill("solid", fgColor="EBF2FA")

THIN_SIDE    = Side(style="thin", color="C8D8E8")
THIN_BORDER  = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

MONEY_FMT    = '#,##0.00'
HEADERS      = ["Nombre", "Categoría", "Unidad", "Stock", "P. Compra", "P. Venta"]
COL_WIDTHS   = [40, 20, 14, 10, 14, 14]

wb = openpyxl.Workbook()
wb.remove(wb.active)

def write_sheet(ws, rows, tab_color="1E3A5F"):
    ws.sheet_properties.tabColor = tab_color

    # Encabezados
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    # Datos
    for row_n, prod in enumerate(rows, start=2):
        row_fill = ALT_FILL if (row_n % 2 == 0) else PatternFill("solid", fgColor="FFFFFF")

        values = [
            prod["nombre"],
            prod["categoria"],
            prod["unidad"],
            prod["stock"],
            prod["precio_compra"],
            prod["precio_venta"],
        ]
        aligns = [DATA_ALIGN_L, DATA_ALIGN_C, DATA_ALIGN_C, DATA_ALIGN_R, DATA_ALIGN_R, DATA_ALIGN_R]
        fmts   = [None, None, None, '#,##0.0', MONEY_FMT, MONEY_FMT]

        for col_idx, (val, aln, fmt) in enumerate(zip(values, aligns, fmts), start=1):
            cell = ws.cell(row=row_n, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.alignment = aln
            cell.border = THIN_BORDER
            cell.fill = row_fill
            if fmt:
                cell.number_format = fmt

        ws.row_dimensions[row_n].height = 16

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"

CAT_COLORS = {
    "JUGUETES":       "E74C3C",
    "ASEO Y LIMPIEZA":"3498DB",
    "ACCESORIOS":     "9B59B6",
    "ESCOLARES":      "F39C12",
    "PRODUCTOS":      "27AE60",
    "CATALOGO":       "16A085",
}

# Hojas por categoría
for cat, prods in by_cat.items():
    sheet_name = cat[:31]
    ws = wb.create_sheet(title=sheet_name)
    color = CAT_COLORS.get(cat, "607D8B")
    write_sheet(ws, prods, tab_color=color)
    print(f"  - Hoja '{sheet_name}': {len(prods)} productos")

# Hoja resumen
ws_all = wb.create_sheet(title="TODOS", index=0)
write_sheet(ws_all, products, tab_color="2C3E50")
print(f"  - Hoja 'TODOS': {len(products)} productos")

wb.save(OUT_PATH)
print(f"\n✅ Proceso exitoso. Archivo final listo para importar:\n   {OUT_PATH}")
